import json
import logging
import os
import platform
import re
import sys  # Replaced os._exit with sys.exit
import threading
import tkinter as tk
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, scrolledtext, ttk
from typing import Set, Optional, Tuple, Callable

import pandas as pd
from azure.ai.documentintelligence import DocumentIntelligenceClient
from azure.core.credentials import AzureKeyCredential
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


##############################################################################
# 1) Setup Logging
##############################################################################

class GuiHandler(logging.Handler):
    """
    Custom logging handler to output logs to a Tkinter ScrolledText widget.
    """

    def __init__(self, text_widget: scrolledtext.ScrolledText):
        super().__init__()
        self.text_widget = text_widget

    def emit(self, record):
        msg = self.format(record)

        def append():
            self.text_widget.insert(tk.END, msg + "\n")
            self.text_widget.see(tk.END)

        self.text_widget.after(0, append)


def setup_logging(gui_text_widget: scrolledtext.ScrolledText):
    """
    Configures logging for the application.

    Args:
        gui_text_widget (ScrolledText): The text widget to display logs.
    """
    logger = logging.getLogger()
    logger.setLevel(logging.DEBUG)  # Set to DEBUG for detailed logs

    # File Handler
    file_handler = logging.FileHandler('processing.log')
    file_handler.setLevel(logging.DEBUG)
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)

    # GUI Handler
    gui_handler = GuiHandler(gui_text_widget)
    gui_handler.setLevel(logging.DEBUG)
    gui_handler.setFormatter(formatter)
    logger.addHandler(gui_handler)


##############################################################################
# 2) Load Stop Words
##############################################################################

def load_stop_words(file_path: str) -> Set[str]:
    """
    Loads stop words from a specified file.

    Args:
        file_path (str): Path to the stop words file.

    Returns:
        Set[str]: A set of stop words.
    """
    try:
        with open(file_path, "r", encoding="utf-8") as file_handle:
            stop_words = {line.strip().lower() for line in file_handle if line.strip()}
            logging.info(f"Loaded {len(stop_words)} stop words from {file_path}")
            return stop_words
    except FileNotFoundError:
        logging.error(f"Stop words file not found: {file_path}")
        messagebox.showerror("Error", f"Stop words file not found: {file_path}")
        return set()
    except Exception as e:
        logging.error(f"Error loading stop words: {e}")
        messagebox.showerror("Error", f"Error loading stop words: {e}")
        return set()


STOP_WORDS_FILE_PATH = "text_files/stop_words.txt"
STOP_WORDS = load_stop_words(STOP_WORDS_FILE_PATH)


##############################################################################
# 3) Credentials Handling
##############################################################################

def read_credentials(credentials_file: Path) -> Tuple[Optional[str], Optional[str]]:
    """
    Reads Azure credentials from a credential file.

    Args:
        credentials_file (Path): Path to the credentials.txt file.

    Returns:
        Tuple[Optional[str], Optional[str]]: AZURE_ENDPOINT and AZURE_KEY if found, else (None, None).
    """
    try:
        with open(credentials_file, "r", encoding="utf-8") as f:
            lines = f.readlines()
        creds = {}
        for line in lines:
            if '=' in line:
                key, value = line.strip().split('=', 1)
                creds[key.strip()] = value.strip()
        azure_endpoint = creds.get("AZURE_ENDPOINT")
        azure_key = creds.get("AZURE_KEY")
        if not azure_endpoint or not azure_key:
            logging.error("AZURE_ENDPOINT or AZURE_KEY not found in credentials.txt")
            return None, None
        logging.info("Successfully read Azure credentials from credentials.txt")
        return azure_endpoint, azure_key
    except FileNotFoundError:
        logging.error(f"Credentials file not found: {credentials_file}")
        messagebox.showerror("Error", f"Credentials file not found: {credentials_file}")
        return None, None
    except Exception as e:
        logging.error(f"Error reading credentials file: {e}")
        messagebox.showerror("Error", f"Error reading credentials file: {e}")
        return None, None


##############################################################################
# 4) JSON to Text Transformation Functions
##############################################################################

def remove_spaces_between_asterisks(text: str) -> str:
    """
    Removes all spaces within asterisks in the given text.

    Example:
        "* An AC - F N - 0507*" -> "*AAC-FN-0507*"

    Args:
        text (str): The input text to process.

    Returns:
        str: The text with spaces removed between asterisks.
    """
    pattern = re.compile(r"\*(.*?)\*")

    def replacer(match):
        inside = match.group(1)
        inside_no_spaces = inside.replace(" ", "")
        return f"*{inside_no_spaces}*"

    return pattern.sub(replacer, text)


def remove_extra_dots(text: str) -> str:
    """
    Removes any sequence of two or more consecutive dots from the text.

    Args:
        text (str): The input text to process.

    Returns:
        str: The text with extra dots removed.
    """
    return re.sub(r"\.{2,}", '', text)


def remove_blank_or_single_dot_lines(text: str) -> str:
    """
    Removes lines that are completely empty or contain only a single dot.

    Args:
        text (str): The input text to process.

    Returns:
        str: The cleaned text with specified lines removed.
    """
    lines = text.splitlines()
    cleaned = [
        line for line in lines
        if line.strip() and line.strip() != "."
    ]
    return "\n".join(cleaned)


def split_financials_and_aacfn(line: str) -> list:
    """
    Splits a line containing both 'AAC - Financials' and 'AAC-FN-XXXX' into separate lines.

    Args:
        line (str): The input line to split.

    Returns:
        list: A list of split lines.
    """
    pattern = re.compile(r"^(.*?\bAAC\s*-\s*Financials\b)\s+(AAC-FN-\S+)(.*)$")
    match = pattern.match(line)
    if match:
        part1 = match.group(1).strip()  # "AAC - Financials"
        part2 = match.group(2).strip()  # "AAC-FN-XXXX"
        part3 = match.group(3).strip()  # Leftover text after AAC-FN-XXXX
        results = [part1, part2]
        if part3:
            results.append(part3)
        return results
    return [line]


def split_aacfn_and_text(line: str) -> list:
    """
    Splits a line containing 'AAC-FN-XXXX' and leftover text into separate lines.

    Args:
        line (str): The input line to split.

    Returns:
        list: A list of split lines.
    """
    pattern = re.compile(r"(AAC-FN-\S+)(\s+)(.+)")
    parts = []
    current = line
    while True:
        match = pattern.search(current)
        if not match:
            parts.append(current)
            break
        before = current[:match.start()].strip()
        if before:
            parts.append(before)
        aacfn = match.group(1)
        parts.append(aacfn)
        leftover = match.group(3).strip()
        current = leftover
    return parts


def fix_financials_and_aacfn(lines: list) -> list:
    """
    Processes lines to ensure proper separation of 'AAC - Financials' and 'AAC-FN-XXXX'.

    Args:
        lines (list): List of input lines.

    Returns:
        list: List of processed and split lines.
    """
    fixed = []
    for line in lines:
        chunks = []
        for piece in split_financials_and_aacfn(line):
            elements = split_aacfn_and_text(piece)
            chunks.extend(elements)
        fixed.extend(chunks)
    return fixed


def is_date_line(line: str) -> bool:
    """
    Determines if a line contains a date in the format "Month DD, YYYY".

    Args:
        line (str): The input line to check.

    Returns:
        bool: True if the line contains a date, False otherwise.
    """
    return bool(re.search(r"[A-Za-z]+\s+\d{1,2},\s*\d{4}", line))


def reorder_lines(lines: list) -> list:
    """
    Reorders lines to ensure that content above '*AAC-FN-XXXX*' is moved appropriately.

    Args:
        lines (list): List of input lines.

    Returns:
        list: List of reordered lines.
    """
    star_line_index = None
    for i, line in enumerate(lines):
        if re.match(r"^\*AAC-FN-\S+\*$", line.strip()):
            star_line_index = i
            break

    if star_line_index is None:
        return lines

    lines_above = lines[:star_line_index]
    lines_main = lines[star_line_index:]
    lines_main = fix_financials_and_aacfn(lines_main)

    final_lines = []
    inserted_above = False
    for i, line in enumerate(lines_main):
        final_lines.append(line)
        if re.match(r"^\*AAC-FN-\S+\*$", line.strip()):
            if i + 1 < len(lines_main):
                next_line = lines_main[i + 1]
                if re.match(r"^AAC-FN-\S+$", next_line.strip()):
                    final_lines.append(next_line)
                    if i + 2 < len(lines_main) and is_date_line(lines_main[i + 2]):
                        final_lines.append(lines_main[i + 2])
                        for ab_line in lines_above:
                            final_lines.append(ab_line)
                        inserted_above = True
                        final_lines.extend(lines_main[i + 3:])
                        break
                    else:
                        for ab_line in lines_above:
                            final_lines.append(ab_line)
                        inserted_above = True
                        final_lines.extend(lines_main[i + 2:])
                        break
    if not inserted_above:
        final_lines.extend(lines_above)

    return final_lines


def process_json_content(content: str) -> str:
    """
    Cleans and reorders JSON content.

    Args:
        content (str): Raw content from JSON.

    Returns:
        str: Processed content.
    """
    content = remove_spaces_between_asterisks(content)
    content = remove_extra_dots(content)
    content = remove_blank_or_single_dot_lines(content)

    lines = content.splitlines()
    lines = reorder_lines(lines)
    return "\n".join(lines)


def extract_and_save_content(json_file: Path, output_folder: Path) -> str:
    """
    Extracts content from a JSON file, processes it, and saves it as a text file.

    Args:
        json_file (Path): Path to the JSON file.
        output_folder (Path): Directory to save the text file.

    Returns:
        str: Path to the saved text file or empty string if failed.
    """
    try:
        with open(json_file, "r", encoding="utf-8") as f:
            data = json.load(f)
        logging.debug(f"Opened JSON file: {json_file}")
    except json.JSONDecodeError as e:
        logging.error(f"Error decoding JSON file {json_file}: {e}")
        return ""
    except FileNotFoundError:
        logging.error(f"JSON file not found: {json_file}")
        return ""
    except Exception as e:
        logging.error(f"Unexpected error opening JSON file {json_file}: {e}")
        return ""

    raw_content = data.get("content", "")
    if not raw_content:
        logging.warning(f"No 'content' field found in JSON file: {json_file}")
    final_output = process_json_content(raw_content)

    base_name = json_file.stem
    output_file = output_folder / f"{base_name}.txt"
    output_folder.mkdir(parents=True, exist_ok=True)
    try:
        with open(output_file, "w", encoding="utf-8") as f:
            f.write(final_output)
        logging.info(f"Saved processed text: {output_file}")
    except Exception as e:
        logging.error(f"Error writing to text file {output_file}: {e}")
        return ""

    return str(output_file)


##############################################################################
# 5) Text File Parsing for Excel Fields
##############################################################################

def parse_text_file(txt_path: str) -> Tuple[str, str, str, str, str]:
    """
    Parses a text file to extract required fields for Excel.

    Args:
        txt_path (str): Path to the text file.

    Returns:
        Tuple[str, str, str, str, str]: Extracted fields (BOX NAME, ID NUMBER, DATE, TEXT, FILE NAME).
    """
    box_name = os.path.basename(os.path.dirname(txt_path))
    file_name = os.path.basename(txt_path)

    try:
        with open(txt_path, "r", encoding="utf-8") as f:
            lines = [line.rstrip("\n") for line in f]
        logging.debug(f"Opened text file: {txt_path}")
    except Exception as e:
        logging.error(f"Error reading text file {txt_path}: {e}")
        return box_name, "NULL", "", "", file_name

    # Extract ID NUMBER
    id_number = "NULL"
    star_pattern = re.compile(r"\*([^*]+)\*")
    for line in lines:
        match = star_pattern.search(line)
        if match:
            id_number = match.group(1).strip()
            logging.debug(f"Extracted ID: {id_number}")
            break

    # Extract DATE
    date_line = ""
    date_regex = re.compile(r"[A-Za-z]+\s+\d{1,2},\s*\d{4}")
    for line in lines:
        if date_regex.search(line):
            date_line = line.strip()
            logging.debug(f"Extracted Date: {date_line}")
            break

    # Extract TEXT
    text_after = []
    if date_line:
        try:
            idx = lines.index(date_line)
            if idx < len(lines) - 1:
                text_after = lines[idx + 1:]
                logging.debug(f"Found text after date at line {idx + 1}")
        except ValueError:
            text_after = []
            logging.warning(f"Date line not found in lines: {txt_path}")
    else:
        fin_idx = None
        for i, ln in enumerate(lines):
            if "AAC - Financials" in ln:
                fin_idx = i
                logging.debug(f"Found 'AAC - Financials' at line {i}")
                break
        if fin_idx is not None and fin_idx < len(lines) - 1:
            text_after = lines[fin_idx + 1:]
            logging.debug(f"Found text after 'AAC - Financials' at line {fin_idx + 1}")
        else:
            text_after = lines
            logging.debug(f"No specific marker found, using entire text for extraction.")

    cleansed_text_after = [
        ln for ln in text_after
        if not re.match(r"^AAC-FN-\S+$", ln.strip())
    ]

    text_after_str = "\n".join(cleansed_text_after).strip()
    text_after_str = re.sub(r"\bAAC-FN-\S+\b", '', text_after_str).strip()

    # Remove '£' symbol
    text_after_str = text_after_str.replace('£', '')

    # Prefix TEXT field with an apostrophe to ensure it's treated as text
    if text_after_str and not text_after_str.startswith("'"):
        text_after_str = f"'{text_after_str}"

    logging.debug(f"Extracted Text: {text_after_str}")

    return box_name, id_number, date_line, text_after_str, file_name


##############################################################################
# 6) Excel Styling Functions
##############################################################################

def excel_worksheet(ws):
    """
    Applies NASA's formatting standards to the Excel worksheet.

    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet to format.
    """
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Apply styles to header
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Apply borders and alignment to all cells
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        adjusted_width = length + 2
        column_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Enable text wrapping for TEXT column (assumed to be column 4)
    for row in ws.iter_rows(min_row=2, max_col=4, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


##############################################################################
# 7) OCR Processing Functions
##############################################################################

def initialize_client(azure_endpoint: str, azure_key: str) -> DocumentIntelligenceClient:
    """
    Initializes the Azure Document Intelligence client.

    Args:
        azure_endpoint (str): The Azure endpoint URL.
        azure_key (str): The Azure key.

    Returns:
        DocumentIntelligenceClient: The initialized client.
    """
    try:
        credential = AzureKeyCredential(azure_key)
        client = DocumentIntelligenceClient(endpoint=azure_endpoint, credential=credential)
        logging.info("Initialized Azure Document Intelligence client.")
        return client
    except Exception as e:
        logging.error(f"Failed to initialize Azure client: {e}")
        messagebox.showerror("Error", f"Failed to initialize Azure client: {e}")
        raise


def process_file(file_path: Path, output_folder: Path, client: DocumentIntelligenceClient) -> str:
    """
    Processes a single PDF/JPEG file using Azure OCR, saves JSON and filtered text.

    Args:
        file_path (Path): Path to the file to process.
        output_folder (Path): Directory to save output files.
        client (DocumentIntelligenceClient): Azure Document Intelligence client.

    Returns:
        str: Summary message of the processing result.
    """
    extension = file_path.suffix.lower()
    base_name = file_path.stem

    json_path_out = output_folder / f"{base_name}.json"
    text_path_out = output_folder / f"{base_name}_filtered.txt"

    SUPPORTED_TYPES = {".pdf": "application/pdf", ".jpg": "image/jpeg", ".jpeg": "image/jpeg"}

    if extension not in SUPPORTED_TYPES:
        logging.info(f"Skipped unsupported file type: {file_path}")
        return f"Skipped (unsupported type): {file_path}\n"

    content_type = SUPPORTED_TYPES[extension]

    try:
        with open(file_path, "rb") as f_in:
            logging.debug(f"Processing file: {file_path}")
            poller = client.begin_analyze_document(
                model_id="prebuilt-read",
                analyze_request=f_in,
                content_type=content_type
            )
            analyze_result = poller.result()

        # Save JSON
        with open(json_path_out, "w", encoding="utf-8") as f_json:
            json.dump(analyze_result.as_dict(), f_json, indent=4, ensure_ascii=False)
        logging.info(f"Saved JSON: {json_path_out}")

        # Save filtered text
        with open(text_path_out, "w", encoding="utf-8") as f_txt:
            for page in analyze_result.pages:
                for line in page.lines:
                    content = line.content
                    tokens = content.split()
                    filtered_tokens = [t for t in tokens if t.lower() not in STOP_WORDS]
                    f_txt.write(" ".join(filtered_tokens) + "\n")
        logging.info(f"Saved filtered text: {text_path_out}")

        return f"Processed: {file_path}\n"

    except Exception as exc:
        error_message = f"Failed to process {file_path}: {exc}"
        logging.error(error_message)
        return error_message + "\n"


##############################################################################
# 8) Main Processing Function
##############################################################################

def handle_folder_upload(input_folder: Path, output_folder: Path, log_callback: Callable[[str], None] = None) -> str:
    """
    Processes all PDF/JPEG files in the input folder to create JSON and text files,
    then parses these JSON files to extract data and compile into an Excel report.

    Args:
        input_folder (Path): Directory containing PDF/JPEG files.
        output_folder (Path): Directory to save JSON, text files, and Excel report.
        log_callback (Callable[[str], None], optional): Function to log messages to the GUI.

    Returns:
        str: Summary of processing results.
    """
    all_rows = []
    processed_files = 0
    skipped_files = 0

    # Read credentials from credentials.txt
    credentials_file = Path("text_files/credentials.txt")
    azure_endpoint, azure_key = read_credentials(credentials_file)
    if not azure_endpoint or not azure_key:
        summary = "Azure credentials could not be read. Please check credentials.txt."
        logging.error(summary)
        if log_callback:
            log_callback(summary)
        return summary

    # Initialize Azure client
    try:
        client = initialize_client(azure_endpoint, azure_key)
    except Exception as e:
        summary = f"Initialization failed: {e}"
        logging.error(summary)
        if log_callback:
            log_callback(summary)
        return summary

    # Step 1: Process PDF/JPEG files to create JSON and text files
    supported_extensions = [".pdf", ".jpg", ".jpeg"]
    pdf_jpeg_files = list(input_folder.rglob("*"))
    pdf_jpeg_files = [f for f in pdf_jpeg_files if f.is_file() and f.suffix.lower() in supported_extensions]

    logging.info(f"Found {len(pdf_jpeg_files)} supported files in input folder.")
    if log_callback:
        log_callback(f"Found {len(pdf_jpeg_files)} supported files to process.")

    if not pdf_jpeg_files:
        summary = "No supported PDF/JPEG files found in the input folder."
        logging.warning(summary)
        if log_callback:
            log_callback(summary)
        return summary

    for file_path in pdf_jpeg_files:
        result = process_file(file_path, output_folder, client)
        if log_callback:
            log_callback(result.strip())
        if result.startswith("Processed"):
            processed_files += 1
        else:
            skipped_files += 1

    # Step 2: Process JSON files to extract data for Excel
    json_files = [f for f in output_folder.glob("*.json")]

    logging.info(f"Found {len(json_files)} JSON files in output folder for data extraction.")
    if log_callback:
        log_callback(f"Found {len(json_files)} JSON files for data extraction.")

    if not json_files:
        summary = "No JSON files were created during processing."
        logging.warning(summary)
        if log_callback:
            log_callback(summary)
        return summary

    for json_file in json_files:
        txt_path = extract_and_save_content(json_file, output_folder)
        if not txt_path:
            skipped_files += 1
            if log_callback:
                log_callback(f"Skipped file due to errors: {json_file.name}")
            continue

        row_data = parse_text_file(txt_path)
        if not any(row_data[1:4]):
            # If ID, Date, and Text are all empty
            logging.warning(f"No data extracted from text file: {txt_path}")
            if log_callback:
                log_callback(f"No data extracted from text file: {json_file.name}")
            skipped_files += 1
            continue

        # Create hyperlink for the text file
        if row_data[4] == "":
            file_hyperlink = "NULL"
        else:
            abs_txt_path = os.path.abspath(txt_path)
            if platform.system() == 'Windows':
                file_url = f'file:///{abs_txt_path.replace("\\", "/")}'
            else:
                file_url = f'file://{abs_txt_path}'
            file_hyperlink = f'=HYPERLINK("{file_url}", "{row_data[4]}")'

        row_with_link = list(row_data)
        row_with_link[4] = file_hyperlink
        all_rows.append(row_with_link)

    # Step 3: Compile data into Excel
    if not all_rows:
        summary = "No data was extracted from any JSON files."
        logging.info(summary)
        if log_callback:
            log_callback(summary)
        return summary

    # Create DataFrame
    df = pd.DataFrame(
        all_rows,
        columns=["BOX NAME", "ID NUMBER", "DATE", "TEXT", "FILE NAME"]
    )

    # Sanitize DataFrame to prevent formulas in TEXT field
    df["TEXT"] = df["TEXT"].apply(
        lambda x: f"'{x}" if isinstance(x, str) and x.startswith('=') else x
    )

    # Remove '£' from all string fields
    string_columns = df.select_dtypes(include=['object']).columns
    for col in string_columns:
        df[col] = df[col].str.replace('£', '', regex=False)

    # Create Excel Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Write header
    for col_num, column_title in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_num, value=column_title)

    # Write data rows
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Apply formatting
    excel_worksheet(ws)

    # Define an Excel output path with timestamp to prevent overwriting
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_output = output_folder / f"output_{timestamp}.xlsx"

    # Save the workbook
    try:
        wb.save(excel_output)
        summary = (
            f"Processed {processed_files} files. "
            f"Skipped {skipped_files} files due to errors.\n"
            f"Excel file saved to: {excel_output}"
        )
        logging.info(summary)
        if log_callback:
            log_callback(summary)
    except Exception as e:
        summary = f"Error saving Excel file: {e}"
        logging.error(summary)
        if log_callback:
            log_callback(summary)

    return summary


##############################################################################
# 9) GUI Implementation
##############################################################################

def start_gui() -> None:
    """
    Initializes and runs the GUI for the OCR & Text Filter Tool.
    """
    selected_input_folder: Optional[Path] = None
    selected_output_folder: Optional[Path] = None

    def select_input_folder() -> None:
        nonlocal selected_input_folder
        folder = filedialog.askdirectory(title="Choose a folder containing PDF/JPEG files to process")
        if folder:
            selected_input_folder = Path(folder)
            input_folder_label.config(text=f"Input folder: {selected_input_folder.name}")
            logging.info(f"Selected input folder: {selected_input_folder}")
        else:
            selected_input_folder = None
            input_folder_label.config(text="No input folder selected.")
            logging.info("No input folder selected.")

    def select_output_folder() -> None:
        nonlocal selected_output_folder
        folder = filedialog.askdirectory(title="Choose a folder to save the output files")
        if folder:
            selected_output_folder = Path(folder)
            output_folder_label.config(text=f"Output folder: {selected_output_folder.name}")
            logging.info(f"Selected output folder: {selected_output_folder}")
        else:
            selected_output_folder = None
            output_folder_label.config(text="No output folder selected.")
            logging.info("No output folder selected.")

    def run_processing() -> None:
        if not selected_input_folder or not selected_output_folder:
            messagebox.showwarning("Folders Missing", "Please select both input and output folders before running.")
            logging.warning("Run attempted without selecting both input and output folders.")
            return

        # Disable the Run button to prevent multiple clicks
        run_button.config(state='disabled')
        status_label.config(text="Processing...")
        progress_bar.start()
        root_window.update_idletasks()

        # Clear the output text box
        output_text_box.delete("1.0", tk.END)

        def log_callback(message: str):
            output_text_box.insert(tk.END, message + "\n")
            output_text_box.see(tk.END)
            logging.info(message)

        def processing_thread():
            try:
                summary = handle_folder_upload(selected_input_folder, selected_output_folder, log_callback)
                progress_bar.stop()
                status_label.config(text="Processing complete.")
                messagebox.showinfo("Success", f"Processing complete.\n{summary}")
            except Exception as e:
                logging.error(f"Processing failed: {e}")
                messagebox.showerror("Error", f"An error occurred during processing:\n{e}")
                status_label.config(text="Processing failed.")
            finally:
                run_button.config(state='normal')

        threading.Thread(target=processing_thread).start()

    # Initialize the main window
    root_window = tk.Tk()
    root_window.title("OCR & Text Filter Tool")
    root_window.geometry("900x700")
    root_window.configure(bg="#f0f0f0")

    # Header Frame
    header_frame = tk.Frame(root_window, bg="#4a90e2", height=60)
    header_frame.pack(fill='x')
    title_label = tk.Label(
        header_frame,
        text="OCR & Text Filter Tool",
        font=("Helvetica", 24, "bold"),
        fg="white",
        bg="#4a90e2"
    )
    title_label.pack(pady=20)

    # Main Frame
    main_frame = tk.Frame(root_window, bg="#f0f0f0")
    main_frame.pack(fill='both', expand=True, padx=20, pady=20)

    # Button Frame
    button_frame = tk.Frame(main_frame, bg="#f0f0f0")
    button_frame.pack(pady=10)

    # Folder Selection Buttons
    input_folder_btn = tk.Button(
        button_frame,
        text="Select Input Folder",
        command=select_input_folder,
        font=("Helvetica", 12),
        bg="#4a90e2",
        fg="white",
        padx=10,
        pady=5
    )
    input_folder_btn.grid(row=0, column=0, padx=10)

    output_folder_btn = tk.Button(
        button_frame,
        text="Select Output Folder",
        command=select_output_folder,
        font=("Helvetica", 12),
        bg="#4a90e2",
        fg="white",
        padx=10,
        pady=5
    )
    output_folder_btn.grid(row=0, column=1, padx=10)

    run_button = tk.Button(
        button_frame,
        text="Run",
        command=run_processing,
        font=("Helvetica", 12),
        bg="#4a90e2",
        fg="white",
        padx=10,
        pady=5
    )
    run_button.grid(row=0, column=2, padx=10)

    # Folder Path Labels
    input_folder_label = tk.Label(
        button_frame,
        text="No input folder selected",
        font=("Helvetica", 12),
        bg="#f0f0f0"
    )
    input_folder_label.grid(row=1, column=0, pady=5)

    output_folder_label = tk.Label(
        button_frame,
        text="No output folder selected",
        font=("Helvetica", 12),
        bg="#f0f0f0"
    )
    output_folder_label.grid(row=1, column=1, pady=5)

    # Output Text Area for Logs
    output_text_box = scrolledtext.ScrolledText(
        main_frame,
        wrap='word',
        width=100,
        height=25,
        font=("Courier", 10)
    )
    output_text_box.pack(pady=10)

    # Setup logging with GUI handler
    setup_logging(output_text_box)

    # Status Label
    status_label = tk.Label(
        main_frame,
        text="",
        font=("Helvetica", 10),
        bg="#f0f0f0",
        fg="#4a90e2"
    )
    status_label.pack(pady=5)

    # Footer Frame with Progress Bar
    footer_frame = tk.Frame(main_frame, bg="#f0f0f0", height=40)
    footer_frame.pack(pady=10, fill='x')
    progress_bar = ttk.Progressbar(
        footer_frame,
        orient="horizontal",
        mode="indeterminate",
        length=400
    )
    progress_bar.pack(pady=10)

    root_window.mainloop()


##############################################################################
# 10) Entry Point
##############################################################################

if __name__ == "__main__":
    # Initialize a temporary Tkinter root to handle potential errors
    temp_root = tk.Tk()
    temp_root.withdraw()  # Hide the main window

    # Read Azure credentials from credentials.txt
    credentials_file = Path("text_files/credentials.txt")
    AZURE_ENDPOINT, AZURE_KEY = read_credentials(credentials_file)

    if not AZURE_ENDPOINT or not AZURE_KEY:
        messagebox.showerror("Configuration Error", "Azure credentials are not properly set in credentials.txt.")
        logging.error("Azure credentials not found or incomplete in credentials.txt.")
        temp_root.destroy()
        sys.exit(1)  # Use sys.exit instead of os._exit
    else:
        temp_root.destroy()
        start_gui()
