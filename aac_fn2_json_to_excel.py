import json
import logging
import os
import platform
import re
import threading
import tkinter as tk
from datetime import datetime
from tkinter import filedialog, messagebox, ttk, scrolledtext

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

##############################################################################
# 1) Setup Logging
##############################################################################

# Configure logging to write to a file with timestamps
logging.basicConfig(
    filename='processing.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)


##############################################################################
# 2) JSON to Text Transformation Functions
##############################################################################

def remove_spaces_between_asterisks(text: str) -> str:
    """
    Removes all spaces within asterisks in the given text.

    Example:
        "* An AC - F N - 0507*" -> "*AAC-FN-0507*"

    Args:
        text (str): The input text to process.

    Returns:
        str: The text with spaces removed between asterisks.
    """
    pattern = re.compile(r"\*(.*?)\*")

    def replacer(match):
        inside = match.group(1)
        inside_no_spaces = inside.replace(" ", "")
        return f"*{inside_no_spaces}*"

    return pattern.sub(replacer, text)


def remove_extra_dots(text: str) -> str:
    """
    Removes any sequence of two or more consecutive dots from the text.

    Args:
        text (str): The input text to process.

    Returns:
        str: The text with extra dots removed.
    """
    return re.sub(r"\.{2,}", '', text)


def remove_blank_or_single_dot_lines(text: str) -> str:
    """
    Removes lines that are completely empty or contain only a single dot.

    Args:
        text (str): The input text to process.

    Returns:
        str: The cleaned text with specified lines removed.
    """
    lines = text.splitlines()
    cleaned = [
        line for line in lines
        if line.strip() and line.strip() != "."
    ]
    return "\n".join(cleaned)


def split_financials_and_aacfn(line: str) -> list:
    """
    Splits a line containing both 'AAC - Financials' and 'AAC-FN-XXXX' into separate lines.

    Args:
        line (str): The input line to split.

    Returns:
        list: A list of split lines.
    """
    pattern = re.compile(r"^(.*?\bAAC\s*-\s*Financials\b)\s+(AAC-FN-\S+)(.*)$")
    match = pattern.match(line)
    if match:
        part1 = match.group(1).strip()  # "AAC - Financials"
        part2 = match.group(2).strip()  # "AAC-FN-XXXX"
        part3 = match.group(3).strip()  # Leftover text after AAC-FN-XXXX
        results = [part1, part2]
        if part3:
            results.append(part3)
        return results
    return [line]


def split_aacfn_and_text(line: str) -> list:
    """
    Splits a line containing 'AAC-FN-XXXX' and leftover text into separate lines.

    Args:
        line (str): The input line to split.

    Returns:
        list: A list of split lines.
    """
    pattern = re.compile(r"(AAC-FN-\S+)(\s+)(.+)")
    parts = []
    current = line
    while True:
        match = pattern.search(current)
        if not match:
            parts.append(current)
            break
        before = current[:match.start()].strip()
        if before:
            parts.append(before)
        aacfn = match.group(1)
        parts.append(aacfn)
        leftover = match.group(3).strip()
        current = leftover
    return parts


def fix_financials_and_aacfn(lines: list) -> list:
    """
    Processes lines to ensure proper separation of 'AAC - Financials' and 'AAC-FN-XXXX'.

    Args:
        lines (list): List of input lines.

    Returns:
        list: List of processed and split lines.
    """
    fixed = []
    for line in lines:
        chunks = []
        for piece in split_financials_and_aacfn(line):
            elements = split_aacfn_and_text(piece)
            chunks.extend(elements)
        fixed.extend(chunks)
    return fixed


def is_date_line(line: str) -> bool:
    """
    Determines if a line contains a date in the format "Month DD, YYYY".

    Args:
        line (str): The input line to check.

    Returns:
        bool: True if the line contains a date, False otherwise.
    """
    return bool(re.search(r"[A-Za-z]+\s+\d{1,2},\s*\d{4}", line))


def reorder_lines(lines: list) -> list:
    """
    Reorders lines to ensure that content above '*AAC-FN-XXXX*' is moved appropriately.

    Args:
        lines (list): List of input lines.

    Returns:
        list: List of reordered lines.
    """
    star_line_index = None
    for i, line in enumerate(lines):
        if re.match(r"^\*AAC-FN-\S+\*$", line.strip()):
            star_line_index = i
            break

    if star_line_index is None:
        return lines

    lines_above = lines[:star_line_index]
    lines_main = lines[star_line_index:]
    lines_main = fix_financials_and_aacfn(lines_main)

    final_lines = []
    inserted_above = False
    for i, line in enumerate(lines_main):
        final_lines.append(line)
        if re.match(r"^\*AAC-FN-\S+\*$", line.strip()):
            if i + 1 < len(lines_main):
                next_line = lines_main[i + 1]
                if re.match(r"^AAC-FN-\S+$", next_line.strip()):
                    final_lines.append(next_line)
                    if i + 2 < len(lines_main) and is_date_line(lines_main[i + 2]):
                        final_lines.append(lines_main[i + 2])
                        for ab_line in lines_above:
                            final_lines.append(ab_line)
                        inserted_above = True
                        final_lines.extend(lines_main[i + 3:])
                        break
                    else:
                        for ab_line in lines_above:
                            final_lines.append(ab_line)
                        inserted_above = True
                        final_lines.extend(lines_main[i + 2:])
                        break
    if not inserted_above:
        final_lines.extend(lines_above)

    return final_lines


def process_json_content(content: str) -> str:
    """
    Cleans and reorders JSON content.

    Args:
        content (str): Raw content from JSON.

    Returns:
        str: Processed content.
    """
    content = remove_spaces_between_asterisks(content)
    content = remove_extra_dots(content)
    content = remove_blank_or_single_dot_lines(content)

    lines = content.splitlines()
    lines = reorder_lines(lines)
    return "\n".join(lines)


def extract_and_save_content(json_file: str, output_folder: str) -> str:
    """
    Extracts content from a JSON file, processes it, and saves it as a text file.

    Args:
        json_file (str): Path to the JSON file.
        Output_folder (str): Directory to save the text file.

    Returns:
        str: Path to the saved text file or empty string if failed.
        :param json_file:
        :param output_folder:
    """
    try:
        with open(json_file, "r", encoding="utf-8") as f:
            data = json.load(f)
    except json.JSONDecodeError as e:
        logging.error(f"Error decoding JSON file {json_file}: {e}")
        return ""

    raw_content = data.get("content", "")
    final_output = process_json_content(raw_content)

    base_name = os.path.splitext(os.path.basename(json_file))[0]
    output_file = os.path.join(output_folder, f"{base_name}.txt")
    os.makedirs(output_folder, exist_ok=True)
    try:
        with open(output_file, "w", encoding="utf-8") as f:
            f.write(final_output)
    except Exception as e:
        logging.error(f"Error writing to text file {output_file}: {e}")
        return ""

    logging.info(f"Saved: {output_file}")
    return output_file


##############################################################################
# 3) Text File Parsing for Excel Fields
##############################################################################

def parse_text_file(txt_path: str) -> tuple:
    """
    Parses a text file to extract required fields for Excel.

    Args:
        txt_path (str): Path to the text file.

    Returns:
        tuple: Extracted fields (BOX NAME, ID NUMBER, DATE, TEXT, FILE NAME).
    """
    box_name = os.path.basename(os.path.dirname(txt_path))
    file_name = os.path.basename(txt_path)

    try:
        with open(txt_path, "r", encoding="utf-8") as f:
            lines = [line.rstrip("\n") for line in f]
    except Exception as e:
        logging.error(f"Error reading text file {txt_path}: {e}")
        return box_name, "NULL", "", "", file_name

    # Extract ID NUMBER
    id_number = "NULL"
    star_pattern = re.compile(r"\*([^*]+)\*")
    for line in lines:
        match = star_pattern.search(line)
        if match:
            id_number = match.group(1).strip()
            break

    # Extract DATE
    date_line = ""
    date_regex = re.compile(r"[A-Za-z]+\s+\d{1,2},\s*\d{4}")
    for line in lines:
        if date_regex.search(line):
            date_line = line.strip()
            break

    # Extract TEXT
    text_after = []
    if date_line:
        try:
            idx = lines.index(date_line)
            if idx < len(lines) - 1:
                text_after = lines[idx + 1:]
        except ValueError:
            text_after = []
    else:
        fin_idx = None
        for i, ln in enumerate(lines):
            if "AAC - Financials" in ln:
                fin_idx = i
                break
        if fin_idx is not None and fin_idx < len(lines) - 1:
            text_after = lines[fin_idx + 1:]
        else:
            text_after = lines

    cleansed_text_after = [
        ln for ln in text_after
        if not re.match(r"^AAC-FN-\S+$", ln.strip())
    ]

    text_after_str = "\n".join(cleansed_text_after).strip()
    text_after_str = re.sub(r"\bAAC-FN-\S+\b", '', text_after_str).strip()

    # Remove '£' symbol
    text_after_str = text_after_str.replace('£', '')

    # Prefix TEXT field with an apostrophe to ensure it's treated as text
    if text_after_str and not text_after_str.startswith("'"):
        text_after_str = f"'{text_after_str}"

    return box_name, id_number, date_line, text_after_str, file_name


##############################################################################
# 4) Excel Styling Functions
##############################################################################

def excel_worksheet(ws):
    """
    Applies NASA's formatting standards to the Excel worksheet.

    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet to format.
    """
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Apply styles to header
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Apply borders and alignment to all cells
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        adjusted_width = length + 2
        column_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Enable text wrapping for TEXT column (assumed to be column 4)
    for row in ws.iter_rows(min_row=2, max_col=4, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


##############################################################################
# 5) Main Processing Function
##############################################################################

def handle_folder_upload(input_folder: str, output_folder: str, log_callback=None) -> str:
    """
    Processes all JSON files in the input folder, converts them to text,
    parses the text files, and compiles the results into a formatted Excel file
    with hyperlinked file names.

    Args:
        input_folder (str): Directory containing JSON files.
        output_folder (str): Directory to save text files and the Excel report.
        log_callback (callable, optional): Function to log messages to the GUI.

    Returns:
        str: Summary of processing results.
    """
    all_rows = []
    processed_files = 0
    skipped_files = 0

    json_files = [f for f in os.listdir(input_folder) if f.lower().endswith(".json")]

    for filename in json_files:
        json_path = os.path.join(input_folder, filename)
        if log_callback:
            log_callback(f"Processing file: {filename}")
        txt_path = extract_and_save_content(json_path, output_folder)
        if not txt_path:
            skipped_files += 1
            if log_callback:
                log_callback(f"Skipped file due to errors: {filename}")
            continue

        row_data = parse_text_file(txt_path)
        if row_data[4] == "":
            file_hyperlink = "NULL"
        else:
            abs_txt_path = os.path.abspath(txt_path)
            if platform.system() == 'Windows':
                file_url = f'file:///{abs_txt_path.replace("\\", "/")}'
            else:
                file_url = f'file://{abs_txt_path}'
            file_hyperlink = f'=HYPERLINK("{file_url}", "{row_data[4]}")'

        row_with_link = list(row_data)
        row_with_link[4] = file_hyperlink
        all_rows.append(row_with_link)
        processed_files += 1
        if log_callback:
            log_callback(f"Processed file: {filename}")

    if not all_rows:
        return "No JSON files were processed."

    # Create DataFrame
    df = pd.DataFrame(
        all_rows,
        columns=["BOX NAME", "ID NUMBER", "DATE", "TEXT", "FILE NAME"]
    )

    # Sanitize DataFrame to prevent formulas in TEXT field
    df["TEXT"] = df["TEXT"].apply(
        lambda x: f"'{x}" if isinstance(x, str) and x.startswith('=') else x
    )

    # Remove '£' from all string fields
    string_columns = df.select_dtypes(include=['object']).columns
    for col in string_columns:
        df[col] = df[col].str.replace('£', '', regex=False)

    # Create Excel Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Write header without assigning to 'cell'
    for col_num, column_title in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_num, value=column_title)

    # Write data rows
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Apply formatting
    excel_worksheet(ws)

    # Define an Excel output path with timestamp to prevent overwriting
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_output = os.path.join(output_folder, f"output_{timestamp}.xlsx")

    # Save the workbook
    try:
        wb.save(excel_output)
        summary = (
            f"Processed {processed_files} files. "
            f"Skipped {skipped_files} files due to errors.\n"
            f"Excel file saved to: {excel_output}"
        )
        logging.info(summary)
        if log_callback:
            log_callback(summary)
    except Exception as e:
        summary = f"Error saving Excel file: {e}"
        logging.error(summary)
        if log_callback:
            log_callback(summary)

    return summary


##############################################################################
# 6) GUI Implementation
##############################################################################

def start_gui() -> None:
    """
    Initializes and runs the GUI for the OCR & Text Filter Tool.
    """
    selected_input_folder: str = ""
    selected_output_folder: str = ""

    def select_input_folder() -> None:
        """
        Opens a dialog to select the input folder containing JSON files.
        """
        nonlocal selected_input_folder
        folder = filedialog.askdirectory(
            title="Choose a folder containing JSON files to process"
        )
        if folder:
            selected_input_folder = folder
            input_folder_label.config(
                text=f"Input folder: {os.path.basename(selected_input_folder)}"
            )
            logging.info(f"Selected input folder: {selected_input_folder}")
        else:
            selected_input_folder = ""
            input_folder_label.config(text="No input folder selected.")
            logging.info("No input folder selected.")

    def select_output_folder() -> None:
        """
        Opens a dialog to select the output folder for saving results.
        """
        nonlocal selected_output_folder
        folder = filedialog.askdirectory(
            title="Choose a folder to save the output files"
        )
        if folder:
            selected_output_folder = folder
            output_folder_label.config(
                text=f"Output folder: {os.path.basename(selected_output_folder)}"
            )
            logging.info(f"Selected output folder: {selected_output_folder}")
        else:
            selected_output_folder = ""
            output_folder_label.config(text="No output folder selected.")
            logging.info("No output folder selected.")

    def run_processing() -> None:
        """
        Initiates the processing of JSON files when the Run button is clicked.
        """
        if not selected_input_folder or not selected_output_folder:
            messagebox.showwarning(
                "Folders missing",
                "Please select both input and output folders before running."
            )
            logging.warning("Run attempted without selecting both input and output folders.")
            return

        # Disable the Run button to prevent multiple clicks
        run_button.config(state=tk.DISABLED)
        status_label.config(text="Processing...")
        progress_bar.start()
        root_window.update_idletasks()

        # Start processing in a separate thread to keep GUI responsive
        processing_thread = threading.Thread(target=process_files)
        processing_thread.start()

    def process_files():
        """
        Handles the file processing in a separate thread.
        """

        def log_callback(message: str) -> None:
            """
            Updates the output text box with log messages.

            Args:
                message (str): The message to display.
            """
            output_text_box.insert(tk.END, message + "\n")
            output_text_box.see(tk.END)

        result_text = handle_folder_upload(
            selected_input_folder,
            selected_output_folder,
            log_callback
        )
        progress_bar.stop()
        status_label.config(text="Processing complete.")
        messagebox.showinfo("Success", f"Processing complete.\n{result_text}")
        run_button.config(state=tk.NORMAL)

    # Initialize the main window
    root_window = tk.Tk()
    root_window.title("OCR & Text Filter Tool (Parallel)")
    root_window.geometry("900x700")
    root_window.configure(bg="#f0f0f0")

    # Main Frame
    main_frame = tk.Frame(root_window, bg="#f0f0f0")
    main_frame.pack(expand=True, fill="both")   
    #main_frame.grid(row=1, column=0, padx=20, pady=20, sticky="nsew")
    #main_frame.grid_columnconfigure(0, weight=1)


    # Header Frame
    header_frame = tk.Frame(main_frame, bg="#4a90e2", height=60)
    #header_frame.grid(row=0, column=0, columnspan=2, sticky="ew")
    header_frame.pack(fill="x")
    title_label = tk.Label(
        header_frame,
        text="OCR & Text Filter Tool (Parallel)",
        font=("Helvetica", 24, "bold"),
        fg="white",
        bg="#4a90e2"
    )
    #title_label.grid(row=0, column=0, padx=10, pady=20, sticky='ew')
    title_label.pack(pady=10)

    # Button Frame
    button_frame = tk.Frame(main_frame, bg="#f0f0f0")
    #button_frame.grid(row=0, column=0, pady=10)
    button_frame.pack(pady=10)

    # Folder Selection Buttons and Path Labels
    input_folder_frame = tk.Frame(button_frame, bg="#f0f0f0")
    input_folder_frame.pack(side="left", padx=10)

    input_folder_btn = tk.Button(
        input_folder_frame,
        text="Select Input Folder",
        command=select_input_folder,
        font=("Helvetica", 12),
        bg="#4a90e2",
        fg="white",
        padx=10,
        pady=5
    )
    #input_folder_btn.grid(row=0, column=0, padx=10)
    input_folder_btn.pack(side="top", padx=10)

    input_folder_label = tk.Label(
        input_folder_frame,
        text="No input folder selected",
        font=("Helvetica", 12),
        bg="#f0f0f0"
    )
    #input_folder_label.grid(row=1, column=0, pady=5)
    input_folder_label.pack(side='top',pady=5) 

    output_folder_frame = tk.Frame(button_frame, bg="#f0f0f0")
    output_folder_frame.pack(side="left", padx=10)

    output_folder_btn = tk.Button(
        output_folder_frame,
        text="Select Output Folder",
        command=select_output_folder,
        font=("Helvetica", 12),
        bg="#4a90e2",
        fg="white",
        padx=10,
        pady=5
    )
    #output_folder_btn.grid(row=0, column=1, padx=10)
    output_folder_btn.pack(side="top", padx=10)

    output_folder_label = tk.Label(
        output_folder_frame,
        text="No output folder selected",
        font=("Helvetica", 12),
        bg="#f0f0f0"
    )
    #output_folder_label.grid(row=1, column=1, pady=5)
    output_folder_label.pack(side='top',pady=5)

    run_button_frame = tk.Frame(button_frame, bg="#f0f0f0") 
    run_button_frame.pack(side="left", padx=10)

    # Run Button
    run_button = tk.Button(
        run_button_frame,
        text="Run",
        command=run_processing,
        font=("Helvetica", 12),
        bg="#4a90e2",
        fg="white",
        padx=20,
        pady=10
    )
    #run_button.grid(row=0, column=2, padx=10)
    run_button.pack(side="left", padx=10)

    # Output Text Area
    output_text_box = scrolledtext.ScrolledText(
        main_frame,
        wrap=tk.WORD,
        width=100,
        height=25,
        font=("Courier", 10)
    )
    #output_text_box.grid(row=1, column=0, pady=10)
    output_text_box.pack(pady=10)

    # Status Label
    status_label = tk.Label(
        main_frame,
        text="",
        font=("Helvetica", 10),
        bg="#f0f0f0",
        fg="#4a90e2"
    )
    #status_label.grid(row=2, column=0, pady=5)
    status_label.pack(pady=5)   

    # Footer Frame with Progress Bar
    footer_frame = tk.Frame(main_frame, bg="#f0f0f0", height=40)
    #footer_frame.grid(row=3, column=0, columnspan=2, pady=10, sticky="ew")
    #footer_frame.grid_columnconfigure(0, weight=1)
    footer_frame.pack(pady=10)  

    progress_bar = ttk.Progressbar(
        footer_frame,
        orient="horizontal",
        mode="indeterminate",
        length=400
    )
    #progress_bar.grid(row=0, column=0, pady=10)
    progress_bar.pack(pady=10)  

    root_window.mainloop()


##############################################################################
# 7) Entry Point
##############################################################################

if __name__ == "__main__":
    start_gui()
